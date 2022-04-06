import openpyxl
from pathlib import Path
from openpyxl import Workbook


class Cell:

    def __init__(self, pos: str):
        self.line = int(pos[1:])
        self.row = ord(pos[0])


class Group:

    def __init__(self, name: str, week1: str, week2: str):
        self.name = name
        self.week1 = week1
        self.week2 = week2


def add_table_to_all(main: list, aux: list):
    pass


def get_xlsx(name: str) -> Workbook:
    file = Path(f'data\\{name}.xlsx')
    table = openpyxl.load_workbook(file, data_only=True)
    return table


def gen_line(arr: list) -> str:
    res = ''
    for i in range(len(arr)):
        if 0 <= i <= 2:
            res += arr[i] + '<br>'
        else:
            if arr[i] != '':
                res += f'<a href="{arr[i]}">Ссылка</a>'
    return res


def read_shorts(sheet) -> dict:
    shorts = {}
    start, end = Cell('J2'), Cell(f'N{int(sheet["B20"].value) + 1}')
    for i in range(start.line, end.line + 1):
        index_key = ''
        index_list = []
        for j in range(start.row, end.row + 1):
            line = str(sheet[f'{chr(j)}{i}'].value)

            if j == start.row:
                index_key = line
                continue
            if line == 'None':
                index_list.append('')
                continue
            if j == start.row+1:
                line += '<br>'

            index_list.append(line)
        shorts[index_key] = gen_line(index_list)
    return shorts


def read_times(sheet) -> dict:
    times = {}
    start, end = Cell('A2'), Cell('C8')
    for i in range(start.line, end.line + 1):
        index_key = ''
        index_value = ''
        for j in range(start.row, end.row + 1):
            line = str(sheet[f'{chr(j)}{i}'].value)
            if ':' not in line:
                index_key = line
                continue
            index_value += line + '<br>'
        times[index_key] = index_value
    return times


def read_table_weeks(sheet, shorts: dict, times: dict) -> Group:

    days = []
    start, end = Cell('B2'), Cell('H2')
    for i in range(start.line, end.line + 1):
        for j in range(start.row, end.row + 1):
            days.append(sheet[f'{chr(j)}{i}'].value)

    week1 = '<table border="1" cellpadding="10" cellspacing="0" rules="all" align="center" ><tr><td style="width:90px;"></td>'
    for i in days:
        week1 += f'<td>{i}</td>'
    week1 += '</tr>'
    start, end = Cell('B3'), Cell('H9')
    for i in range(start.line, end.line + 1):
        week1 += f'<tr><td>{i-start.line+1} пара<br>{times[str(i-start.line+1)]}</td>'
        for j in range(start.row, end.row + 1):
            value = str(sheet[f'{chr(j)}{i}'].value)
            if value != 'None':
                line = shorts.get(value)
            else:
                line = ''
            week1 += f'<td>{line}</td>'
        week1 += '</tr>'
    week1 += '</table>'

    week2 = '<table border="1" cellpadding="10" cellspacing="0" rules="all" align="center" ><tr><td style="width:90px;"></td>'
    for i in days:
        week2 += f'<td>{i}</td>'
    week2 += '</tr>'
    start, end = Cell('B12'), Cell('H18')
    for i in range(start.line, end.line + 1):
        week2 += f'<tr><td>{i-start.line+1} пара<br>{times[str(i-start.line+1)]}</td>'
        for j in range(start.row, end.row + 1):
            value = str(sheet[f'{chr(j)}{i}'].value)
            if value != 'None':
                line = shorts.get(value)
            else:
                line = ''
            week2 += f'<td>{line}</td>'
        week2 += '</tr>'
    week2 += '</table>'

    return Group(sheet.title, week1, week2)


def read_sheet(sheet, timesheet) -> Group:
    times = read_times(timesheet)
    shorts = read_shorts(sheet)
    return read_table_weeks(sheet, shorts, times)


def init(names: list):
    groups = []
    for i in names:
        if i != '':
            book = get_xlsx(i)
            timesheet = ''
            for j in book.sheetnames:
                if 'Час' in j:
                    timesheet = book[j]
            for j in book.sheetnames:
                if 'Час' not in j:
                    groups.append(read_sheet(book[j], timesheet))
    return groups

